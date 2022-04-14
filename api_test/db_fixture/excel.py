# -*- coding:utf-8 -*-#
import os
import sys

from db_fixture.func import fun

current_dir = os.path.dirname(__file__)
parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
sys.path.append(parent_dir)

import xlrd3, ast, requests, json
from openpyxl import load_workbook


# 读取参数Excel

class db():
    # apiurl = "http://192.192.192.53:9001"  # 后台IP
    apiurl = "http://192.192.192.53:8401" #APP

    def __init__(self):

        self.method = ''  # 请求方法
        self.url = ''  # 接口地址
        self.header = []# 前期header
        self.data = []  # 请求参数
        self.except_code = ''  # 预期code
        self.except_result = ''  # 预期结果
        self.actual_result = ''  # 实际结果
        self.actual_code = ''  # 实际code
        self.test_res = ''  # 测试结果

    # @classmethod
    def readExcel(self, file_path, i):
        """
        读取Excel接口用例需要的参数
        :param file_path: 文件地址
        :param i: 用例的条数
        :return: 请求方法，请求地址，请求参数，请求header
        """

        # 检查文件地址是否有误
        try:
            book = xlrd3.open_workbook(file_path)  # 打开Excel
        except Exception:
            print('路径不在或文件错误')

        else:
            sheet = book.sheet_by_index(0)  # 从第一个sheet开始取
            # sheet = book.sheet_by_name('') # 按表格的名字取
            self.rows = sheet.nrows  # 取这个的所有行数
            # 获取case行
            case_list = sheet.row_values(i)
            # print(case_list)

            self.method = case_list[4]  # 获取请求方法
            self.url = self.apiurl + case_list[5]  # 获取接口地址
            self.except_code = case_list[8]  # 获取预期code
            self.except_result = case_list[9]  # 获取预期结果
            if len(case_list[6]) == 0: # 判断header是否为空，如果为空，则返回none
                self.header = None
            else:
                self.header = ast.literal_eval(case_list[6])
            # cls.header = ast.literal_eval(case_list[6])
            if len(case_list[7]) == 0: # 判断bady是否为空 如果为空，则返回none
                self.data =  None
            else:
                self.data = ast.literal_eval(case_list[7])  # 获取接口参数,将字符串转化为字典

            # print(cls.method,cls.url,cls.data,cls.header)
            return self.method,self.url,self.data,self.header


    def send_request(self):
        """
        调用请求接口
        """
        # print(self.url)
        print(self.method, self.url, self.data,self.header)
        self.actual_result = fun.run_method(self.method, self.url, self.data,self.header)
        print(self.actual_result)
        return self.actual_result

    def decide(self):
        """
        断言，并将结果写入excel
        :return: 断言结果
        """
        # 先判断接口响应是否成功
        try:
            self.actual_code = json.loads(self.actual_result)['code']
        except:
            self.test_res = 'Http请求出错'
            print('Http请求出错')
            return self.actual_result
        else:

            if self.actual_code == self.except_code:
                self.test_res = '验证通过'
                return self.test_res
            else:
                self.test_res = '验证失败'
                return self.test_res


    def write_excel(cls, i):
        """
        将接口返回的结果和断言结果存入Excel
        :param result: 实际结果
        :param i: 用例序号
        :param test_res: 测试结果
        :return:
        """
        # 将返回结果存入Excel表中result_append
        wb = load_workbook(file_path)
        ws = wb['Sheet1']
        local = 'k%s' % (i + 1)
        local_test = 'L%s' % (i + 1)
        ws[local] = cls.actual_result
        ws[local_test] = cls.test_res
        wb.save(file_path)

    def running(self, file_path):
        for i in range(1, 13):
            self.readExcel(file_path, i)
            self.send_request()
            self.decide()
            self.write_excel(i)
            print(i)

    def running_sl(self, file_path):
        """
        运行单个用例
        :param file_path:
        :return:
        """
        i = 6
        self.readExcel(file_path, i)
        self.send_request()
        self.decide()
        self.write_excel(i)



    def __iter__(self):
        """
        迭代器，遍历case

        :return:
        """



    def read_yaml(self):
        """
        读取yaml文件
        :return:
        """

if __name__ == '__main__':
    file_path = r'C:/Users/DELL/Desktop/yongli.xlsx'
    a = db()
    a.running_sl(file_path)
